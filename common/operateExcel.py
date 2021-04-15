import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, colors
from config import configPath


class Variable():
    # excel中每列数据对应的索引
    id = 1
    testCase = 3
    priority = 4
    Account = 5
    Password = 6
    Expect = 7
    isExecute = 8
    TestResult = 9

    # test_result = 2

# 操作Excel数据
class OperateExcel:
    def __init__(self):
        self.file_path = configPath.excel_path
        self.wb = load_workbook(self.file_path)

    # 获取一列数据
    def get_column_value(self, sheet_name, col_num):
        sh = self.wb[sheet_name]
        case = []
        for i in range(1, sh.max_row + 1):
            value = sh.cell(i, col_num).value
            case.append(value)
        return case

    # 获取id的序列(从而确定这一行数据)
    def get_rowId_index(self, sheet_name, id):
        idx = Variable.id
        allIdList = self.get_column_value(sheet_name, col_num=idx)
        if id not in allIdList:
            raise TypeError('编号：{},不存在!'.format(id))
        else:
            idIndex = allIdList.index(id)
            return idIndex

    # 获取一行数据
    def get_row_value(self, sheet_name, row_num):
        sh = self.wb[sheet_name]
        case = []
        for i in range(1, sh.max_column + 1):
            value = sh.cell(row_num, i).value
            case.append(value)
        return case

    # 获取单元格数据
    def get_cell_value(self, sheet_name, col_num, row_num):
        sh = self.wb[sheet_name]
        value = sh.cell(row_num, col_num).value
        return value

    # 获取是否执行用例字段
    def get_is_execute(self,sheet_name,id):
        sh = self.wb[sheet_name]
        isExecute = sh.cell(self.get_rowId_index(sheet_name,id) + 1,Variable.isExecute).value
        if isExecute.lower() == "no":
            flag = False
        elif isExecute.lower() == "yes":
            flag = True
        else:
            raise TypeError("用户编号为：{} 的 ‘ IsExecute ’ 字段有误！允许类型：yes、no ".format(id))
        return flag

    # 获取测试用例字段
    def get_test_case(self,sheet_name,id):
        sh = self.wb[sheet_name]
        testCase = sh.cell(self.get_rowId_index(sheet_name,id) + 1,Variable.testCase).value
        return testCase

    # 获取账号密码字段
    def get_account_data(self,sheet_name,id):
        sh = self.wb[sheet_name]
        data = []
        account = sh.cell(self.get_rowId_index(sheet_name,id) + 1,Variable.Account).value
        data.append(account)
        password = sh.cell(self.get_rowId_index(sheet_name,id) + 1,Variable.Password).value
        data.append(password)
        for i in range(len(data)):
            if data[i] is None:
                data[i] = ''
            return data

    # 写入单元格数据
    def write_result(self, sheet_name, id, value):
        """
        @param sheet_name: Sheet表名称
        @param id:用例编号
        @param value: fail or pass
        """

        sh = self.wb[sheet_name]
        # 设置时间格式
        curTime = datetime.datetime.now().strftime('%Y%m%d %H:%M:%S')
        # 横,纵坐标
        y = self.get_rowId_index(sheet_name, id) + 1
        x = Variable.TestResult
        # 写入
        sh.cell(y, x).value = curTime + "->" + value
        # 设置字体样式
        if value.lower() == 'fail':
            v = sh.cell(y, x)
            v.font = Font(name='Arial', color=colors.RED, size=10)
        elif value.lower() == 'pass':
            v = sh.cell(y, x)
            v.font = Font(name='Arial', color="00B050",size=10)
        else:
            raise TypeError('请传入fail,pass,不区分大小写')
        # 保存
        print('保存数据,执行成功')
        self.wb.save(self.file_path)





