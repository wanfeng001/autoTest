import os
import time
from datetime import date

import pytest
from openpyxl import load_workbook
from common import configpath

file_path = configpath.excel_path
file1_path = configpath.excel1_path


# 读取Excel数据
class ReadExcel:
    def __init__(self):
        self.file_path = file1_path
        self.wb = load_workbook(self.file_path)

    # 获取某一列数据[不包括首行]
    def get_column_value(self, sheet_name, col_num):
        sh = self.wb[sheet_name]
        case = []
        for i in range(2, sh.max_row + 1):
            value = sh.cell(i, col_num).value
            case.append(value)
        return case

    # 获取某一行数据[不包括首列]
    def get_row_value(self, sheet_name, row_num):
        sh = self.wb[sheet_name]
        case = []
        for i in range(2, sh.max_column + 1):
            value = sh.cell(row_num, i).value
            case.append(value)
        return case

    # 获取单元格数据
    def get_cell_value(self, sheet_name, col_num, row_num):
        sh = self.wb[sheet_name]
        value = sh.cell(row_num, col_num).value
        return value

    # 写入单元格数据
    def write_cell(self, sheet_name, col_num, row_num, value):
        sh = self.wb[sheet_name]
        sh.cell(row_num, col_num).value = value
        self.wb.save(self.file_path)


# 获取执行为Y的数据
def get_data():
    read_excel = ReadExcel()
    case = []
    for idx, value in enumerate(read_excel.get_column_value(sheet_name='登录模块用例', col_num=configpath.isExecute)):
        if value == 'Y':
            data = read_excel.get_row_value(sheet_name='登录模块用例', row_num=idx + 2)
            # 继续添加想要的数据
            username = data[configpath.Account - 2]
            password = data[configpath.Password - 2]
            info = [username, password]
            for i in range(len(info)):
                if info[i] == None:
                    info[i] = ''
            case.append(info)
    return case


if __name__ == '__main__':
    print(get_data())
